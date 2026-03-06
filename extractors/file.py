"""Extract structured text from uploaded files (Excel, CSV, plain text)."""

import io
import pandas as pd
import openpyxl


def extract_excel(content: bytes, filename: str = "") -> dict:
    """
    Parse a multi-sheet Excel file into a text representation for Claude.
    Each sheet becomes a section: sheet name + pipe-delimited rows.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Skip entirely empty rows
            if all(cell is None for cell in row):
                continue
            row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
            rows.append(row_str)

        if rows:
            sections.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))

    text = "\n\n".join(sections)
    title = filename.rsplit(".", 1)[0] if filename else "Uploaded Spreadsheet"

    return {
        "text": text,
        "title": title,
        "date": None,
        "source_type": "spreadsheet",
    }


def extract_csv(content: bytes, filename: str = "") -> dict:
    """Parse a CSV file into pipe-delimited text."""
    try:
        df = pd.read_csv(io.BytesIO(content))
    except Exception:
        df = pd.read_csv(io.BytesIO(content), encoding="latin-1")

    rows = [" | ".join(str(v) for v in df.columns)]
    for _, row in df.iterrows():
        rows.append(" | ".join(str(v) for v in row))

    text = "\n".join(rows)
    title = filename.rsplit(".", 1)[0] if filename else "Uploaded CSV"

    return {
        "text": text,
        "title": title,
        "date": None,
        "source_type": "spreadsheet",
    }


def extract_text(content: bytes, filename: str = "") -> dict:
    """Decode a plain text file."""
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    title = filename.rsplit(".", 1)[0] if filename else "Uploaded Text"
    return {
        "text": text.strip(),
        "title": title,
        "date": None,
        "source_type": "text",
    }


def extract(content: bytes, filename: str) -> dict:
    """Dispatch to the appropriate parser based on file extension."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        return extract_excel(content, filename)
    elif ext == "csv":
        return extract_csv(content, filename)
    else:
        return extract_text(content, filename)
